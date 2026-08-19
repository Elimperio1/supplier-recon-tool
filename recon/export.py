"""Excel workbook builder (BUILD.md §8 App checklist).

The workbook is a second renderer over the same engine result the UI shows.
Rules that matter: amounts are real numbers (so the accountant can sum them),
descriptions/refs are forced to text (``@``) so ``16/03/2026``-shaped descriptions
and long numeric refs are never coerced to dates/scientific notation, the header
row is frozen, and green/red/amber fills flag status at a glance.
"""

from __future__ import annotations

import io
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .engine import (CAT_GREEN, CAT_INVOICES, CAT_PAYMENTS, CrossSettlement,
                     CrossSupplierFinding, EngineResult, SupplierResult)
from .match import (VERDICT_AMBIGUOUS, VERDICT_CONFIDENT, VERDICT_NONE, MatchResult)
from .parse import BankReport, SupplierReport

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
GREEN_FONT = Font(color="006100")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
RED_FONT = Font(color="9C0006")
AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")
AMBER_FONT = Font(color="9C6500")

MONEY_FMT = "#,##0.00"
TEXT_FMT = "@"

CATEGORY_LABEL = {
    CAT_GREEN: "GREEN",
    CAT_PAYMENTS: "PAYMENTS NEEDED",
    CAT_INVOICES: "INVOICES NEEDED",
}
VERDICT_STYLE = {
    VERDICT_CONFIDENT: (GREEN_FILL, GREEN_FONT),
    VERDICT_AMBIGUOUS: (AMBER_FILL, AMBER_FONT),
    VERDICT_NONE: (RED_FILL, RED_FONT),
}


def _rand(cents: Optional[int]) -> Optional[float]:
    return None if cents is None else round(cents / 100.0, 2)


def _header(ws: Worksheet, headers: list[str]) -> None:
    for c, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws: Worksheet, headers: list[str], widths: Optional[dict[int, int]] = None) -> None:
    widths = widths or {}
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(c, 18)


def _put(ws: Worksheet, r: int, c: int, value, *, money=False, text=False):
    cell = ws.cell(row=r, column=c, value=value)
    if money and value is not None:
        cell.number_format = MONEY_FMT
    elif text:
        cell.number_format = TEXT_FMT
    return cell


# ---------------------------------------------------------------------------

def _summary_sheet(ws: Worksheet, engine: EngineResult) -> None:
    headers = ["Supplier", "Closing (R)", "Category", "Integrity", "Notes"]
    _header(ws, headers)
    _autosize(ws, headers, {1: 38, 3: 18, 4: 22, 5: 44})
    r = 2
    for res in sorted(engine.suppliers, key=lambda x: (x.category != CAT_PAYMENTS, x.name.lower())):
        _put(ws, r, 1, res.name, text=True)
        _put(ws, r, 2, _rand(res.closing), money=True)
        cat_cell = _put(ws, r, 3, CATEGORY_LABEL.get(res.category, res.category))
        fill, font = (GREEN_FILL, GREEN_FONT) if res.category == CAT_GREEN else (RED_FILL, RED_FONT)
        cat_cell.fill, cat_cell.font = fill, font
        s = res.supplier
        integ = "OK" if s.integrity_ok else f"MISMATCH off by {_rand(s.integrity_delta)}"
        ic = _put(ws, r, 4, integ)
        if not s.integrity_ok:
            ic.fill, ic.font = RED_FILL, RED_FONT
        _put(ws, r, 5, "; ".join(res.notes), text=True)
        r += 1


def _payments_sheet(ws: Worksheet, matches: dict[str, list[MatchResult]],
                    engine: EngineResult) -> None:
    headers = ["Supplier", "Invoice Ref", "Amount (R)", "Verdict", "Bank Account",
               "Bank Date", "Bank Description", "Bank Ref", "Current GL", "Evidence", "Note"]
    _header(ws, headers)
    _autosize(ws, headers, {1: 30, 2: 16, 4: 12, 5: 16, 6: 12, 7: 40, 8: 16, 9: 26, 10: 20, 11: 40})
    r = 2
    for res in engine.by_category(CAT_PAYMENTS):
        if res.bulk:
            _put(ws, r, 1, res.name, text=True)
            _put(ws, r, 4, "BULK")
            _put(ws, r, 11, "; ".join(res.notes), text=True)
            r += 1
            continue
        for m in matches.get(res.name, []):
            fill, font = VERDICT_STYLE.get(m.verdict, (None, None))
            cands = m.candidates or [None]
            for i, cand in enumerate(cands):
                _put(ws, r, 1, res.name if i == 0 else "", text=True)
                _put(ws, r, 2, m.invoice.reference if i == 0 else "", text=True)
                _put(ws, r, 3, _rand(m.amount_cents) if i == 0 else None, money=True)
                vc = _put(ws, r, 4, m.verdict.upper() if i == 0 else "")
                if i == 0 and fill:
                    vc.fill, vc.font = fill, font
                if cand is not None:
                    _put(ws, r, 5, cand.txn.account, text=True)
                    _put(ws, r, 6, cand.txn.date, text=True)
                    _put(ws, r, 7, cand.txn.description, text=True)
                    _put(ws, r, 8, cand.txn.reference, text=True)
                    _put(ws, r, 9, cand.txn.allocation, text=True)
                    _put(ws, r, 10, ", ".join(cand.matched_tokens), text=True)
                if i == 0 and m.note:
                    _put(ws, r, 11, m.note, text=True)
                r += 1


def _invoices_sheet(ws: Worksheet, engine: EngineResult) -> None:
    headers = ["Supplier", "Payment Ref", "Amount (R)", "Date", "Description", "Flag"]
    _header(ws, headers)
    _autosize(ws, headers, {1: 30, 2: 16, 5: 44, 6: 20})
    r = 2
    for res in engine.by_category(CAT_INVOICES):
        for t in res.unmatched_payments:
            _put(ws, r, 1, res.name, text=True)
            _put(ws, r, 2, t.reference, text=True)
            _put(ws, r, 3, _rand(t.debit), money=True)
            _put(ws, r, 4, t.date, text=True)
            _put(ws, r, 5, t.description, text=True)
            _put(ws, r, 6, "request invoice", text=True)
            r += 1


def _cross_sheet(ws: Worksheet, findings: Iterable[CrossSupplierFinding]) -> None:
    headers = ["Kind", "Supplier A", "Supplier B", "Amount (R)", "Evidence"]
    _header(ws, headers)
    _autosize(ws, headers, {1: 18, 2: 32, 3: 32, 5: 30})
    r = 2
    for f in findings:
        _put(ws, r, 1, f.kind, text=True)
        _put(ws, r, 2, f.supplier_a, text=True)
        _put(ws, r, 3, f.supplier_b, text=True)
        _put(ws, r, 4, _rand(f.amount_cents), money=True)
        _put(ws, r, 5, ", ".join(f.evidence), text=True)
        r += 1


SETTLEMENT_LABEL = {"name_linked": "Likely same vendor", "amount_only": "Amount only"}
SETTLEMENT_STYLE = {"name_linked": (GREEN_FILL, GREEN_FONT), "amount_only": (AMBER_FILL, AMBER_FONT)}


def _settlement_sheet(ws: Worksheet, settlements: Iterable[CrossSettlement]) -> None:
    headers = ["Confidence", "Invoice Supplier", "Invoice Ref", "Payment Supplier",
               "Payment Ref", "Amount (R)", "Shared"]
    _header(ws, headers)
    _autosize(ws, headers, {1: 20, 2: 30, 3: 16, 4: 30, 5: 16, 7: 24})
    r = 2
    for s in settlements:
        cc = _put(ws, r, 1, SETTLEMENT_LABEL.get(s.confidence, s.confidence))
        fill, font = SETTLEMENT_STYLE.get(s.confidence, (None, None))
        if fill:
            cc.fill, cc.font = fill, font
        _put(ws, r, 2, s.invoice_supplier, text=True)
        _put(ws, r, 3, s.invoice_ref, text=True)
        _put(ws, r, 4, s.payment_supplier, text=True)
        _put(ws, r, 5, s.payment_ref, text=True)
        _put(ws, r, 6, _rand(s.amount_cents), money=True)
        _put(ws, r, 7, ", ".join(s.evidence), text=True)
        r += 1


def _typos_sheet(ws: Worksheet, engine: EngineResult) -> None:
    headers = ["Supplier", "Invoice Ref", "Invoice (R)", "Payment Ref", "Payment (R)", "Diff (R)"]
    _header(ws, headers)
    _autosize(ws, headers, {1: 30, 2: 16, 4: 16})
    r = 2
    for res in engine.suppliers:
        for tp in res.typos:
            _put(ws, r, 1, res.name, text=True)
            _put(ws, r, 2, tp.invoice.reference, text=True)
            _put(ws, r, 3, _rand(tp.invoice.credit), money=True)
            _put(ws, r, 4, tp.payment.reference, text=True)
            _put(ws, r, 5, _rand(tp.payment.debit), money=True)
            _put(ws, r, 6, _rand(tp.diff_cents), money=True)
            r += 1


def _integrity_sheet(ws: Worksheet, engine: EngineResult, supplier_report: SupplierReport,
                     bank_report: Optional[BankReport]) -> None:
    headers = ["Type", "Name", "Reported (R)", "Recomputed (R)", "Delta (R)", "Detail"]
    _header(ws, headers)
    _autosize(ws, headers, {1: 16, 2: 38, 6: 40})
    r = 2
    for s in supplier_report.integrity_failures:
        _put(ws, r, 1, "supplier", text=True)
        _put(ws, r, 2, s.name, text=True)
        _put(ws, r, 3, _rand(s.closing), money=True)
        _put(ws, r, 4, _rand(s.recomputed_closing), money=True)
        _put(ws, r, 5, _rand(s.integrity_delta), money=True)
        r += 1
    if bank_report:
        for a in bank_report.integrity_failures:
            _put(ws, r, 1, "bank", text=True)
            _put(ws, r, 2, a.name, text=True)
            _put(ws, r, 3, _rand(a.closing), money=True)
            _put(ws, r, 4, _rand(a.recomputed_closing), money=True)
            _put(ws, r, 5, _rand(a.integrity_delta), money=True)
            r += 1
    # Counts block
    _put(ws, r, 1, "count", text=True)
    _put(ws, r, 2, "unrecognized supplier rows", text=True)
    _put(ws, r, 6, str(len(supplier_report.unrecognized)), text=True)
    r += 1
    if bank_report is not None:
        _put(ws, r, 1, "count", text=True)
        _put(ws, r, 2, "unrecognized bank rows", text=True)
        _put(ws, r, 6, str(len(bank_report.unrecognized)), text=True)
        r += 1
    if supplier_report.duplicate_names:
        _put(ws, r, 1, "warn", text=True)
        _put(ws, r, 2, "duplicate supplier names", text=True)
        _put(ws, r, 6, "; ".join(supplier_report.duplicate_names), text=True)
        r += 1


def build_workbook(
    client: str,
    engine: EngineResult,
    matches: dict[str, list[MatchResult]],
    supplier_report: SupplierReport,
    bank_report: Optional[BankReport] = None,
) -> Workbook:
    wb = Workbook()
    _summary_sheet(wb.active, engine)
    wb.active.title = "Summary"
    _payments_sheet(wb.create_sheet("Payments Needed"), matches, engine)
    _invoices_sheet(wb.create_sheet("Invoices Needed"), engine)
    _settlement_sheet(wb.create_sheet("Cross-Account"), engine.settlements)
    _cross_sheet(wb.create_sheet("Cross-Supplier"), engine.cross)
    _typos_sheet(wb.create_sheet("Capture Typos"), engine)
    _integrity_sheet(wb.create_sheet("Integrity"), engine, supplier_report, bank_report)
    return wb


def workbook_bytes(*args, **kwargs) -> bytes:
    buf = io.BytesIO()
    build_workbook(*args, **kwargs).save(buf)
    return buf.getvalue()
